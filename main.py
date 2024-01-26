from gettext import dpgettext
from telegram.ext import CallbackQueryHandler, ApplicationBuilder, ContextTypes, CommandHandler, ConversationHandler, MessageHandler, filters
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, CallbackQuery
import openpyxl
import logging
from typing import Union
import requests
import time
import asyncio
import random
import string
from urllib.parse import urlparse, parse_qs
from datetime import datetime
import stripe

ADMIN_USERS = ["@", "@"]
MOD_USER = ["@", "@", "@"]

NOWPAYMENTS_API_KEY = ""
NOWPAYMENTS_API_ENDPOINT = "https://api.nowpayments.io/v1/invoice"
REQUIRED_AMOUNT = ""
USDT_DECIMALS = ""

async def updateFile():
    global wb, sheet
    wb.close()
    wb = openpyxl.load_workbook('userdata.xlsx')
    sheet = wb['Foglio1']

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

async def require(condition : bool, returnedError : str, update, ctx):
        if(condition):
            return
        else:
            await ctx.bot.send_message(chat_id=update.effective_chat.id, text=returnedError)
            raise Exception(returnedError)
        
def userInsideSheet(userId: int) -> bool:
    for row_num in range(2, sheet.max_row + 1): 
        cell_value = sheet[f'A{row_num}'].value
        if cell_value and str(userId) == str(cell_value):
            print(f"Found userId: {userId} in sheet at row {row_num}")
            return True
        else:
            print(f"Checking userId: {userId} against value in sheet: {cell_value}")
    return False

def userRow(userId: int) -> Union[int, None]:
    if userInsideSheet(userId):
        for row_num in range(2, sheet.max_row + 1): 
            cell_value = sheet[f'A{row_num}'].value
            if cell_value and str(userId) == str(cell_value):
                return row_num
    return None

def pushData(userId : int, column : str, data : str):
    if userInsideSheet(userId):
        sheet[column + str(userRow(userId))] = data
        wb.save('userdata.xlsx')

async def controllo_campi(user_id):
    global sheet
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if str(cell.value) == str(user_id):
                print(f"Found user_id at row {cell.row}")
                columns_to_check = ["C", "D", "E", "F"]
                for col_letter in columns_to_check:
                    value = sheet[f"{col_letter}{cell.row}"].value
                    print(f"Value at {col_letter}{cell.row} is {value}")
                    if value is None or value == "":
                        return False
                return True
    print(f"User_id {user_id} not found in the sheet")
    return False

REGOLAMENTO_TEXT = """

"""

PRIVACYPOLICY_TEXT = """

"""

REGISTRAZIONE_TEXT = """

"""

PAGAMENTO_TEXT = """

"""

stripe.api_key = ""

def create_stripe_checkout_session(amount: int = 100, currency: str = 'eur', description: str = ''):
    try:
        session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{
                'price_data': {
                    'currency': currency,
                    'product_data': {
                        'name': '',
                        'description': description,
                    },
                    'unit_amount': amount,
                },
                'quantity': 1,
            }],
            mode='payment',
            success_url='',
            cancel_url='',
        )
        return session
    except Exception as e:
        print(f"Error Stripe: {str(e)}")
        return None

def create_payment_order(amount: float):
    headers = {
        "x-api-key": NOWPAYMENTS_API_KEY,
        "Content-Type": "application/json"
    }

    order_id = generate_random_order_id()

    data = {
        "price_amount": amount,
        "price_currency": "EUR",
        "pay_currency": "USDTBSC",
        "ipn_callback_url": "YOUR_IPN_CALLBACK_URL",
        "order_id": order_id,
        "order_description": "",
        "success_url": "",
        "cancel_url": ""
    }

    response = requests.post(NOWPAYMENTS_API_ENDPOINT, headers=headers, json=data)

    if 200 <= response.status_code < 300:
        return response.json()
    else:
        print(f"Errore payment order {response.text}")
        return None

def generate_random_order_id(length=10):
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

async def show_user_data_buttons(uuid, ctx):
    await updateFile()
    if not userInsideSheet(uuid):
        await ctx.bot.send_message(chat_id=uuid, text="Usa il comando /accetto per accettare la privacy policy e proseguire con la registrazione!", parse_mode='HTML')
        return
    row = userRow(uuid)
    idtg, privacy, nome_cognome, cod_fiscale, ind_fatturazione, registrazione, data_inizio_abbonamento, data_fine_abbonamento = (
        sheet[f"A{row}"].value, sheet[f"B{row}"].value, sheet[f"C{row}"].value,
        sheet[f"D{row}"].value, sheet[f"E{row}"].value, sheet[f"G{row}"].value,
        sheet[f"L{row}"].value, sheet[f"M{row}"].value
    )

    keyboard = [
        [InlineKeyboardButton(f"Nome e cognome: {nome_cognome}", callback_data="show_nome_cognome")],
        [InlineKeyboardButton(f"Codice fiscale: {cod_fiscale}", callback_data="show_cod_fiscale")],
        [InlineKeyboardButton(f"Indirizzo: {ind_fatturazione}", callback_data="show_ind_fatturazione")],
        [InlineKeyboardButton(f"Registrazione: {registrazione}", callback_data="show_registrazione")],
        [InlineKeyboardButton(f"Inizio abbonamento: {data_inizio_abbonamento}", callback_data="show_inizio_abbonamento")],
        [InlineKeyboardButton(f"Fine abbonamento: {data_fine_abbonamento}", callback_data="show_fine_abbonamento")],
        [InlineKeyboardButton("🔙 Torna al menu", callback_data="menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await ctx.bot.send_message(chat_id=uuid, text="Ecco i tuoi dati:", reply_markup=reply_markup, parse_mode='HTML')

async def handle_callback_query(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query: CallbackQuery = update.callback_query
    query_data = query.data
    uuid = query.from_user.id
    
    if query_data == "dati":
        await show_user_data_buttons(uuid, ctx)
        
    elif query_data == "regolamento":
        keyboard = [[InlineKeyboardButton("🔙 Torna al menu", callback_data="menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await ctx.bot.send_message(
            chat_id=uuid,
            text=REGOLAMENTO_TEXT,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )

    elif query_data == "privacy_policy":
        keyboard = [
            [InlineKeyboardButton("Accetta la privacy policy", callback_data="accetto_privacy")],
            [InlineKeyboardButton("🔙 Torna al menu", callback_data="menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await ctx.bot.send_message(
            chat_id=uuid,
            text=PRIVACYPOLICY_TEXT,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )

    elif query_data == "accetto_privacy":
        await updateFile()
        if userInsideSheet(uuid):
            await ctx.bot.send_message(chat_id=uuid, text="Hai già accettato la privacy policy.")
        else:
            rows = list(sheet.rows)
            await ctx.bot.send_message(chat_id=uuid, text="Grazie per aver accettato la privacy policy, prosegui con la registrazione!")
            sheet[f"A{len(rows)+1}"] = uuid
            sheet[f"B{len(rows)+1}"] = True
            username = query.from_user.username
            sheet[f"F{len(rows)+1}"] = username

            wb.save('userdata.xlsx')

    elif query_data == "registrazione":
        await updateFile()
        if not userInsideSheet(uuid):
            keyboard = [
                [InlineKeyboardButton("Accetta la privacy policy", callback_data="accetto_privacy")],
                [InlineKeyboardButton("🔙 Torna al menu", callback_data="menu")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await ctx.bot.send_message(
                chat_id=uuid,
                text="Devi accettare prima la privacy policy per continuare la registrazione.",
                reply_markup=reply_markup,
                parse_mode='HTML'
            )
        else:
            keyboard_back = [[InlineKeyboardButton("🔙 Torna al menu", callback_data="menu")]]
            reply_markup_back = InlineKeyboardMarkup(keyboard_back)
            await ctx.bot.send_message(
                chat_id=uuid,
                text=REGISTRAZIONE_TEXT,
                reply_markup=reply_markup_back,
                parse_mode='HTML'
            )
    elif query_data == "menu":
        await start(update, ctx)

    elif query_data == "pagamento":
        await updateFile()
        if not userInsideSheet(uuid):
            keyboard = [
                [InlineKeyboardButton("Accetta la privacy policy", callback_data="accetto_privacy")],
                [InlineKeyboardButton("🔙 Torna al menu", callback_data="menu")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await ctx.bot.send_message(
                chat_id=uuid,
                text="Devi accettare prima la privacy policy per continuare la registrazione.",
                reply_markup=reply_markup,
                parse_mode='HTML'
            )
        else:
            keyboard_pay = [
                [InlineKeyboardButton("Crea un ordine e procedi al pagamento", callback_data="start_payment_process")],
                [InlineKeyboardButton("Crea un ordine e procedi al pagamento con carta", callback_data="start_payment_process_stripe")],
                [InlineKeyboardButton("🔙 Torna al menu", callback_data="menu")]
            ]
            reply_markup_pay = InlineKeyboardMarkup(keyboard_pay)
            await ctx.bot.send_message(
                chat_id=uuid,
                text=PAGAMENTO_TEXT,
                reply_markup=reply_markup_pay,
                parse_mode='HTML'
            )
    elif query_data == "menu":
        await start(update, ctx)
    

    elif query_data == "start_payment_process":
        user_id = query.from_user.id
        if not await controllo_campi(user_id): 
            await ctx.bot.send_message(chat_id=uuid, text="Per favore, completa tutti i campi prima di procedere al pagamento.")
            return
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if str(cell.value) == str(user_id):
                    g_cell = sheet[f"G{cell.row}"]
                    g_cell.value = "in corso"
                    payment_order = create_payment_order(REQUIRED_AMOUNT)
                    print(payment_order)
                    if not payment_order:
                        await ctx.bot.send_message(chat_id=uuid, text="Si è verificato un errore durante la creazione dell'ordine di pagamento. Riprova più tardi.")
                        return
                    created_at_iso = payment_order["created_at"]
                    created_at_datetime = datetime.fromisoformat(created_at_iso.replace("Z", "+00:00"))
                    formatted_created_at = created_at_datetime.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    h_cell = sheet[f"H{cell.row}"]
                    i_cell = sheet[f"I{cell.row}"]
                    k_cell = sheet[f"K{cell.row}"]
                    l_cell = sheet[f"L{cell.row}"]
                    h_cell.value = payment_order["id"]
                    i_cell.value = payment_order["invoice_url"]
                    k_cell.value = payment_order["order_id"]
                    l_cell.value = formatted_created_at
                    wb.save('userdata.xlsx')
                    payment_order_url = payment_order["invoice_url"]
                    keyboard = [[InlineKeyboardButton("Effettua il pagamento del tuo ordine", url=payment_order_url)]]
                    reply_markup = InlineKeyboardMarkup(keyboard)
                    await ctx.bot.send_message(chat_id=uuid, text="Clicca sul bottone qui sotto per effettuare il pagamento:", reply_markup=reply_markup)
    
    elif query_data == "start_payment_process_stripe":
        user_id = query.from_user.id
        if not await controllo_campi(user_id):
            await ctx.bot.send_message(chat_id=uuid, text="Per favore, completa tutti i campi prima di procedere al pagamento.")
            return
        session = create_stripe_checkout_session()
        if not session:
            await ctx.bot.send_message(chat_id=uuid, text="Si è verificato un errore durante la creazione della sessione di pagamento. Riprova più tardi.")
            return
        payment_url = session.url
        keyboard = [[InlineKeyboardButton("Effettua il pagamento", url=payment_url)]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await ctx.bot.send_message(chat_id=uuid, text="Clicca sul bottone qui sotto per effettuare il pagamento:", reply_markup=reply_markup)     

    await query.answer()

async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("📜 Regolamento", callback_data="regolamento"),
         InlineKeyboardButton("🔒 Privacy Policy", callback_data="privacy_policy")],
        
        [InlineKeyboardButton("🖊️ Registrazione", callback_data="registrazione"),
         InlineKeyboardButton("📋 Dati", callback_data="dati")],
        
        [InlineKeyboardButton("💳 Pagamento", callback_data="pagamento"),
         InlineKeyboardButton("📞 Supporto", callback_data="supporto")],
        
        [InlineKeyboardButton("📢 Canale Ufficiale", callback_data="canale_ufficiale")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)
    await ctx.bot.send_message(chat_id=update.effective_chat.id, text="<b>Seleziona una delle opzioni:</b>", reply_markup=reply_markup, parse_mode='HTML')

async def nome_cognome(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await updateFile()
    uuid = update.message.from_user['id']
    await require(userInsideSheet(uuid), "Usa il comando /accetto per accettare la privacy policy e proseguire con la registrazione!", update, ctx)
    payment_status = sheet[f"G{userRow(uuid)}"].value
    if payment_status == "in corso":
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Non è possibile cambiare il nome e cognome mentre è in corso un pagamento.")
        return
    elif payment_status == "attivo":
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Il tuo account è già stato attivato, impossibile cambiare il nome e cognome.\n\ncontatta il supporto.")
        return

    if sheet[f"C{userRow(uuid)}"].value == None:
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Imposto il tuo nome e cognome in: " + ''.join(ctx.args))
    else:
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Ho aggiornato il tuo nome e cognome in: " + ''.join(ctx.args))
    pushData(uuid, "C", ''.join(ctx.args))

async def cod_fiscale(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await updateFile()
    uuid = update.message.from_user['id']
    await require(userInsideSheet(uuid), "Usa il comando /accetto per accettare la privacy policy e proseguire con la registrazione!", update, ctx)
    payment_status = sheet[f"G{userRow(uuid)}"].value
    if payment_status == "in corso":
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Non è possibile cambiare il codice fiscale mentre è in corso un pagamento.")
        return
    elif payment_status == "attivo":
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Il tuo account è già stato attivato, impossibile cambiare il codice fiscale.\n\ncontatta il supporto.")
        return

    if sheet[f"D{userRow(uuid)}"].value == None:
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Imposto il tuo codice fiscale in: " + ''.join(ctx.args))
    else:
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Ho aggiornato il tuo codice fiscale in: " + ''.join(ctx.args))
    pushData(uuid, "D", ''.join(ctx.args))

async def ind_fatturazione(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await updateFile()
    uuid = update.message.from_user['id']
    await require(userInsideSheet(uuid), "Usa il comando /accetto per accettare la privacy policy e proseguire con la registrazione!", update, ctx)
    payment_status = sheet[f"G{userRow(uuid)}"].value
    if payment_status == "in corso":
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Non è possibile cambiare l'indirizzo di fatturazione mentre è in corso un pagamento.")
        return
    elif payment_status == "attivo":
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Il tuo account è già stato attivato, impossibile cambiare l'indirizzo di fatturazione.\n\ncontatta il supporto.")
        return

    if sheet[f"E{userRow(uuid)}"].value == None:
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Imposto il tuo indirizzo di fatturazione in: " + ''.join(ctx.args))
    else:
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Ho aggiornato il tuo indirizzo di fatturazione in: " + ''.join(ctx.args))
    pushData(uuid, "E", ''.join(ctx.args))

async def email(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await updateFile()
    uuid = update.message.from_user['id']
    await require(userInsideSheet(uuid), "Usa il comando /accetto per accettare la privacy policy e proseguire con la registrazione!", update, ctx)
    payment_status = sheet[f"G{userRow(uuid)}"].value
    if payment_status == "in corso":
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Non è possibile cambiare l'indirizzo email mentre è in corso un pagamento.")
        return
    elif payment_status == "attivo":
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Il tuo account è già stato attivato, impossibile cambiare l'indirizzo email.\n\ncontatta il supporto.")
        return

    if sheet[f"M{userRow(uuid)}"].value == None:
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Imposto il tuo indirizzo email in: " + ''.join(ctx.args))
    else:
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Ho aggiornato il tuo indirizzo email in: " + ''.join(ctx.args))
    pushData(uuid, "M", ''.join(ctx.args))

async def send_all_users(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    username = update.message.from_user.username
    if "@" + username not in ADMIN_USERS:
        await update.message.reply_text("Non hai il permesso di utilizzare questo comando.")
        return
    await updateFile()
    all_users = []
    for row_num in range(2, sheet.max_row + 1): 
        nome_cognome = sheet[f"C{row_num}"].value
        cod_fiscale = sheet[f"D{row_num}"].value
        ind_fatturazione = sheet[f"E{row_num}"].value
        registrazione = sheet[f"G{row_num}"].value
        if nome_cognome:
            user_info = (
                f"Nome e Cognome: {nome_cognome}, "
                f"Codice Fiscale: {cod_fiscale}, "
                f"Indirizzo Fatturazione: {ind_fatturazione}, "
                f"Registrazione: {registrazione}"
            )
            all_users.append(user_info)
    response = "\n\n".join(all_users)
    await update.message.reply_text(response)

async def checkuser(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    username = "@" + update.message.from_user.username
    if username not in MOD_USER:
        await update.message.reply_text("Non hai il permesso di utilizzare questo comando.")
        return
    if not ctx.args or len(ctx.args) < 1:
        await update.message.reply_text("Per favore, fornisci l'ID del pagamento dopo il comando /checkuser.")
        return
    payment_id = ctx.args[0]
    await updateFile()
    for row_num in range(2, sheet.max_row + 1):
        j_value = sheet[f"J{row_num}"].value
        f_value = sheet[f"F{row_num}"].value
        if j_value == payment_id:
            g_value = sheet[f"G{row_num}"].value
            await update.message.reply_text(f"Registrazione per l'ID di pagamento {payment_id}: {g_value}")
            return
    await update.message.reply_text(f"ID di pagamento {payment_id} non trovato.")

def check_payment_status(payment_id: str):
    url = f"https://api.nowpayments.io/v1/payment/{payment_id}"
    headers = {
        "x-api-key": NOWPAYMENTS_API_KEY
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Errore durante la verifica dello stato del pagamento: {response.text}")
        return None

async def checkpagamento(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await updateFile()
    uuid = update.message.from_user['id']
    if ctx.args and len(ctx.args) > 0:
        payment_id = ctx.args[0]
        print(ctx.args)
        pushData(uuid, "J", ''.join(ctx.args))
        payment_status = check_payment_status(payment_id)
        user_row = None
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if str(cell.value) == str(uuid):
                    user_row = cell.row
                    break
            if user_row:
                break
        if not user_row:
            await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Utente non trovato, effettua prima la registrazione!")
            return
        if payment_status:
            if payment_status.get('payment_status') == 'finished':
                await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Il tuo pagamento è stato confermato!")
                g_cell = sheet[f"G{user_row}"]
                g_cell.value = "attivo"
                wb.save('userdata.xlsx')
            else:
                await ctx.bot.send_message(chat_id=update.effective_chat.id, text=f"Stato del pagamento: {payment_status.get('payment_status')}")
        else:
            await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Errore nella verifica del pagamento.")
    else:
        await ctx.bot.send_message(chat_id=update.effective_chat.id, text="Per favore, fornisci l'ID del pagamento dopo il comando /checkpagamento.")

if __name__ == "__main__":
    wb = openpyxl.load_workbook('userdata.xlsx')
    sheet = wb['Foglio1']
    app = ApplicationBuilder().token('').build()
    app.add_handlers(
        [
            CommandHandler('start', start), 
            CommandHandler("nome_cognome", nome_cognome), 
            CommandHandler("cod_fiscale", cod_fiscale),
            CommandHandler("ind_fatturazione", ind_fatturazione),
            CommandHandler("email", email),
            CommandHandler("invio", send_all_users),
            CommandHandler("checkpagamento", checkpagamento),
            CommandHandler("checkuser", checkuser),
            CallbackQueryHandler(handle_callback_query),
        ]
    )
    app.run_polling()
